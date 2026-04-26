import logging
import json

from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_http_methods

from content.models import Domain, OfferType, Organization

logger = logging.getLogger(__name__)


@require_GET
def import_template(request):
    import io  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415
    from openpyxl.styles import Alignment, Font, PatternFill  # noqa: PLC0415
    from openpyxl.worksheet.datavalidation import DataValidation  # noqa: PLC0415

    offer_type_names = sorted(OfferType.objects.values_list("name", flat=True))
    org_names = sorted(Organization.objects.values_list("name", flat=True))
    domain_names = sorted(Domain.objects.values_list("name", flat=True))
    profiles = ["company", "researcher", "student"]
    iso_countries = [
        "AD", "AE", "AF", "AG", "AI", "AL", "AM", "AO", "AQ", "AR", "AS", "AT", "AU", "AW", "AX",
        "AZ", "BA", "BB", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BL", "BM", "BN", "BO", "BQ",
        "BR", "BS", "BT", "BV", "BW", "BY", "BZ", "CA", "CC", "CD", "CF", "CG", "CH", "CI", "CK",
        "CL", "CM", "CN", "CO", "CR", "CU", "CV", "CW", "CX", "CY", "CZ", "DE", "DJ", "DK", "DM",
        "DO", "DZ", "EC", "EE", "EG", "EH", "ER", "ES", "ET", "FI", "FJ", "FK", "FM", "FO", "FR",
        "GA", "GB", "GD", "GE", "GF", "GG", "GH", "GI", "GL", "GM", "GN", "GP", "GQ", "GR", "GS",
        "GT", "GU", "GW", "GY", "HK", "HM", "HN", "HR", "HT", "HU", "ID", "IE", "IL", "IM", "IN",
        "IO", "IQ", "IR", "IS", "IT", "JE", "JM", "JO", "JP", "KE", "KG", "KH", "KI", "KM", "KN",
        "KP", "KR", "KW", "KY", "KZ", "LA", "LB", "LC", "LI", "LK", "LR", "LS", "LT", "LU", "LV",
        "LY", "MA", "MC", "MD", "ME", "MF", "MG", "MH", "MK", "ML", "MM", "MN", "MO", "MP", "MQ",
        "MR", "MS", "MT", "MU", "MV", "MW", "MX", "MY", "MZ", "NA", "NC", "NE", "NF", "NG", "NI",
        "NL", "NO", "NP", "NR", "NU", "NZ", "OM", "PA", "PE", "PF", "PG", "PH", "PK", "PL", "PM",
        "PN", "PR", "PS", "PT", "PW", "PY", "QA", "RE", "RO", "RS", "RU", "RW", "SA", "SB", "SC",
        "SD", "SE", "SG", "SH", "SI", "SJ", "SK", "SL", "SM", "SN", "SO", "SR", "SS", "ST", "SV",
        "SX", "SY", "SZ", "TC", "TD", "TF", "TG", "TH", "TJ", "TK", "TL", "TM", "TN", "TO", "TR",
        "TT", "TV", "TW", "TZ", "UA", "UG", "UM", "US", "UY", "UZ", "VA", "VC", "VE", "VG", "VI",
        "VN", "VU", "WF", "WS", "YE", "YT", "ZA", "ZM", "ZW",
    ]

    wb = openpyxl.Workbook()

    # Hidden lookups sheet avoids 255-char formula limit for long dropdown lists.
    wl = wb.create_sheet("Lookups")
    wl.sheet_state = "hidden"
    for i, v in enumerate(offer_type_names, 1):
        wl.cell(row=i, column=1, value=v)
    for i, v in enumerate(org_names, 1):
        wl.cell(row=i, column=2, value=v)
    for i, v in enumerate(profiles, 1):
        wl.cell(row=i, column=3, value=v)
    for i, v in enumerate(iso_countries, 1):
        wl.cell(row=i, column=4, value=v)
    for i, v in enumerate(domain_names, 1):
        wl.cell(row=i, column=5, value=v)

    ws = wb.active
    ws.title = "Import"

    HEADERS = [
        "url", "offer_type", "organization", "target_profile", "country",
        "title", "summary",
        "domain_1", "domain_2", "domain_3", "domain_4", "domain_5",
    ]
    REQUIRED = {"url", "offer_type", "organization", "target_profile", "country"}
    COL_WIDTHS = [45, 20, 40, 16, 10, 35, 50, 20, 20, 20, 20, 20]

    fill_req = PatternFill("solid", fgColor="D6EAF8")
    fill_opt = PatternFill("solid", fgColor="EAF7EC")
    font_hdr = Font(bold=True)

    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = font_hdr
        cell.fill = fill_req if header in REQUIRED else fill_opt
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = width

    example = [
        "https://example.com/program",
        offer_type_names[0] if offer_type_names else "training",
        org_names[0] if org_names else "My Organisation",
        "student",
        "IT",
        "Example Title",
        "A short description.",
        domain_names[0] if len(domain_names) > 0 else "",
        domain_names[1] if len(domain_names) > 1 else "",
        "", "", "",
    ]
    for col, value in enumerate(example, 1):
        ws.cell(row=2, column=col, value=value)

    dv_offer_type = DataValidation(
        type="list",
        formula1=f"Lookups!$A$1:$A${len(offer_type_names)}",
        allow_blank=True,
        showDropDown=False,
    )
    dv_org = DataValidation(
        type="list",
        formula1=f"Lookups!$B$1:$B${len(org_names)}",
        allow_blank=True,
        showDropDown=False,
    )
    dv_profile = DataValidation(
        type="list",
        formula1=f"Lookups!$C$1:$C${len(profiles)}",
        allow_blank=True,
        showDropDown=False,
    )
    dv_country = DataValidation(
        type="list",
        formula1=f"Lookups!$D$1:$D${len(iso_countries)}",
        allow_blank=True,
        showDropDown=False,
    )
    domain_formula = f"Lookups!$E$1:$E${len(domain_names)}" if domain_names else '"(no domains)"'
    for col_letter in ["H", "I", "J", "K", "L"]:
        dv = DataValidation(
            type="list",
            formula1=domain_formula,
            allow_blank=True,
            showDropDown=False,
        )
        dv.sqref = f"{col_letter}2:{col_letter}2000"
        ws.add_data_validation(dv)

    ws.add_data_validation(dv_offer_type)
    ws.add_data_validation(dv_org)
    ws.add_data_validation(dv_profile)
    ws.add_data_validation(dv_country)

    dv_offer_type.sqref = "B2:B2000"
    dv_org.sqref = "C2:C2000"
    dv_profile.sqref = "D2:D2000"
    dv_country.sqref = "E2:E2000"

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="oss_import_template.xlsx"'},
    )


@csrf_exempt
@require_http_methods(["POST"])
def import_preview(request):
    from content.ingestion.importer import ImportService  # noqa: PLC0415

    f = request.FILES.get("file")
    if not f:
        return JsonResponse({"error": "No file provided. Send as multipart field 'file'."}, status=400)

    try:
        result = ImportService().preview(f, f.name)
    except (ValueError, ImportError, UnicodeDecodeError):
        logger.warning("Import preview: bad file format", exc_info=True)
        return JsonResponse({"error": "Failed to parse file. Check format and required columns."}, status=400)
    except Exception:
        logger.exception("Import preview: unexpected error")
        return JsonResponse({"error": "An unexpected error occurred. Please try again."}, status=500)

    return JsonResponse(result.to_dict())


@csrf_exempt
@require_http_methods(["POST"])
def import_confirm(request):
    from content.ingestion.importer import ImportService  # noqa: PLC0415

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON body."}, status=400)

    valid_rows = body.get("rows", [])

    if not isinstance(valid_rows, list):
        return JsonResponse({"error": "'rows' must be a list."}, status=400)

    for i, entry in enumerate(valid_rows):
        if not isinstance(entry, dict):
            return JsonResponse({"error": f"Row {i}: expected object, got {type(entry).__name__}."}, status=400)
        if not isinstance(entry.get("data"), dict):
            return JsonResponse({"error": f"Row {i}: missing or invalid 'data' field."}, status=400)
        _REQUIRED_DATA_FIELDS = ("url", "offer_type", "organization", "target_profile", "country")
        for field in _REQUIRED_DATA_FIELDS:
            if not entry["data"].get(field):
                return JsonResponse({"error": f"Row {i}: 'data.{field}' is required."}, status=400)

    try:
        result = ImportService().confirm(valid_rows)
    except Exception:
        logger.exception("Import confirm failed")
        return JsonResponse({"error": "Import failed. Please try again or contact support."}, status=500)

    return JsonResponse(result.to_dict())
